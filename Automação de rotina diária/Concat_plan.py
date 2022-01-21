import pandas as pd
from config import index_stat, all_index, dest_file, dia_atual, ano_atual, mes_atual, dia_semana, func, file, n_file, \
    meses_str, manual_index, aut_index, fatiamento, conversor_dt, conversor_ldt, dias_p_entrega, \
    to_date_time, aut_index_a
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl import Workbook
import time
from DF_LeadTime import df_lead
from DF_Urano import df_urano, qy_ent
from SSW import name_m_atual, search_data
from DF_Veloz import df_veloz

try:
    from DF_Coletados import df_col
    from SSW import name_m_ant
    from DF_Veloz import df_veloz_m
except ImportError:
    pass

start_time = time.time()

itens_comp = []
sheet_names = []

sheet_name = 'Clientes com Agendamento'
df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
sheet_names.append(sheet_name)
itens_comp.append(df)

sheet_name = 'Contato Transportadoras'
df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
sheet_names.append(sheet_name)
itens_comp.append(df)

df_n = pd.read_excel(n_file, engine='openpyxl')

dfs = []

try:
    for dfss in range(1, 13):
        dfss = str(dfss)
        try:
            df = pd.read_excel(file, engine='openpyxl', sheet_name=meses_str[dfss] + '-' + ano_atual)
            df.columns = all_index
        except ValueError:
            continue

        dfs.append(df)
except IndexError:
    pass
print("10% concluído")
try:
    df_n.columns = aut_index
except ValueError:
    df_n.columns = df_n.columns = aut_index_a
for x in df_n['Número']:
    print(x)
df_n = df_n[aut_index_a]
df_n = df_n.dropna()
for x in manual_index:
    df_n[x] = ''

if int(dia_atual) == 1:
    if int(dia_atual) == 2 and dia_semana == 0:
        dfs[int(mes_atual) - 2] = pd.concat([dfs[int(mes_atual) - 2], df_n],
                                            keys=all_index)

    if int(dia_atual) == 3 and dia_semana == 0:
        dfs[int(mes_atual) - 2] = pd.concat([dfs[int(mes_atual) - 2], df_n],
                                            keys=all_index)

    dfs[int(mes_atual) - 2] = pd.concat([dfs[int(mes_atual) - 2], df_n],
                                        keys=all_index)
else:
    print('Incluido notas emitidas do dia anterior')

    dfs[int(mes_atual) - 1] = pd.concat([dfs[int(mes_atual) - 1], df_n],
                                        keys=all_index)

dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].astype(dtype='str')

dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].apply(
    func=lambda val: to_date_time(val))

dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].astype(dtype='str')

dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].apply(
    func=lambda val: conversor_dt(val))
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')
try:
    dfs[int(mes_atual) - 1]['Número'] = dfs[int(mes_atual) - 1]['Número'].apply(func=lambda val: pd.to_numeric(val))
    dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_col, how='left', on='Número')
    dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')
    dfs[int(mes_atual) - 1]['Data-De-Coleta_col'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'] + '!' + \
                                                    dfs[int(mes_atual) - 1]['Data-De-Coleta_col']
    dfs[int(mes_atual) - 1]['Data-De-Coleta_col'] = dfs[int(mes_atual) - 1]['Data-De-Coleta_col'].apply(
        func=lambda val: fatiamento(val))
    dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta_col']
except NameError:
    print('Não há novas atualizações de data de coleta')

if mes_atual != '1':
    try:
        if int(dia_atual) in range(0, 10):
            dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=df_veloz_m, how='left', on='Número')

            # Conversão para str
            dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'].astype(dtype='str')

            # Concatenação das informações
            dfs[int(mes_atual) - 2]['Data_De_Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'] + '!' + \
                                                        dfs[int(mes_atual) - 2]['Data_De_Coleta']

            # Fatiamento da informação
            dfs[int(mes_atual) - 2]['Data_De_Coleta'] = dfs[int(mes_atual) - 2]['Data_De_Coleta'].apply(
                func=lambda val: fatiamento(val))

            # Alocação dentro do DF
            dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data_De_Coleta']
    except NameError:
        # Exceção tratada no if
        pass
print("20% concluído")

# Processo Veloz
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_veloz, how='left', on='Número')
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')
dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].astype(dtype='str')
dfs[int(mes_atual) - 1]['Data_De_Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'] + '!' + \
                                            dfs[int(mes_atual) - 1]['Data_De_Coleta']
dfs[int(mes_atual) - 1]['Data_De_Coleta'] = dfs[int(mes_atual) - 1]['Data_De_Coleta'].apply(
    func=lambda val: fatiamento(val))
dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data_De_Coleta']

# Processo Lead Time
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_lead, how='left',
                                   on=['Cidade-Destinatário', 'Uf', 'Fantasia_Do_Transportador'], copy=False)
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].drop_duplicates(subset=['Número'], keep='first')
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead_Time']
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead-Time'].apply(
    lambda val: pd.to_numeric(val, 'coerce', 'integer'))
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead-Time'].fillna('-')

# Processo SSW
ssw_mes = search_data(name_m_atual)  # DF do mês
df_ssw_col = ssw_mes.query("Ult_Status == ' MERCADORIA RECEBIDA PARA TRANSPORTE'")
df_ssw_ent = ssw_mes.query("Ult_Status == ' MERCADORIA ENTREGUE (01)'")
df_ssw_age = ssw_mes.query("Ult_Status == ' ENTREGA AGENDADA (15)'")
df_ssw_age['Descrição'] = df_ssw_age['Descrição'].apply(func=lambda val: val[19:26] if len(val) > 5 else '')
df_ssw_ent['Data'] = df_ssw_ent['Data'].astype(dtype='str', copy=True)
df_ssw_ent['Data'] = df_ssw_ent['Data'].apply(
    func=lambda val: val[8:12] + '-' + val[5:7] + '-' + val[:4] if len(val) > 3 else val)
ssw_mes = ssw_mes[['Número', 'Data', 'Cidade/Estado', 'Ult_Status']]
ssw_mes['Status'] = ssw_mes['Ult_Status'] + ' - ' + ssw_mes['Cidade/Estado']
dfs[int(mes_atual) - 1]['Agendamento'] = dfs[int(mes_atual) - 1]['Agendamento'].astype(dtype='str', copy=True)
dfs[int(mes_atual) - 1]['D_Entrega'] = dfs[int(mes_atual) - 1]['D_Entrega'].astype(dtype='str', copy=True)
dfs[int(mes_atual) - 1]['D_Entrega'] = dfs[int(mes_atual) - 1]['D_Entrega'].apply(
    func=lambda val: conversor_dt(val))
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_ssw_col, how='left', on='Número')
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')
dfs[int(mes_atual) - 1]['Data'] = dfs[int(mes_atual) - 1]['Data'].astype(dtype='str')
dfs[int(mes_atual) - 1]['Data'] = dfs[int(mes_atual) - 1]['Data'].apply(func=lambda val: conversor_dt(val))
dfs[int(mes_atual) - 1]['Data'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'] + '!' + dfs[int(mes_atual) - 1]['Data']
dfs[int(mes_atual) - 1]['Data'] = dfs[int(mes_atual) - 1]['Data'].apply(func=lambda val: fatiamento(val))
dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data']
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1][all_index]
print("30% concluído")
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_ssw_ent, how='left', on='Número')
dfs[int(mes_atual) - 1].fillna('-', inplace=True)
dfs[int(mes_atual) - 1]['Data'] = dfs[int(mes_atual) - 1]['D_Entrega'] + '!' + dfs[int(mes_atual) - 1]['Data']
dfs[int(mes_atual) - 1]['Data'] = dfs[int(mes_atual) - 1]['Data'].apply(
    func=lambda val: fatiamento(val))
dfs[int(mes_atual) - 1]['D_Entrega'] = dfs[int(mes_atual) - 1]['Data']
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1][all_index]
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=ssw_mes, how='left')
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1][index_stat]
dfs[int(mes_atual) - 1].columns = all_index
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_ssw_age, how='left', on='Número')
dfs[int(mes_atual) - 1]['Descrição'] = dfs[int(mes_atual) - 1]['Descrição'].fillna('')
dfs[int(mes_atual) - 1]['Agendamento'] = dfs[int(mes_atual) - 1]['Agendamento'].apply(
    func=lambda val: val[8:10] + val[5:7] + val[2:4] if len(val) > 4 else '')
dfs[int(mes_atual) - 1]['Descrição'] = dfs[int(mes_atual) - 1]['Agendamento'] + '!' + dfs[int(mes_atual) - 1][
    'Descrição']
dfs[int(mes_atual) - 1]['Descrição'] = dfs[int(mes_atual) - 1]['Descrição'].replace(to_replace='!', value='',
                                                                                    regex=True)
dfs[int(mes_atual) - 1]['Descrição'] = dfs[int(mes_atual) - 1]['Descrição'].astype(dtype='str', copy=True)
dfs[int(mes_atual) - 1]['Descrição'] = dfs[int(mes_atual) - 1]['Descrição'].apply(
    func=lambda val: val[-6:-4] + '-' + val[-4:-2] + '-' + '20' + val[-2:] if len(val) > 1 else '-')
dfs[int(mes_atual) - 1]['Agendamento'] = dfs[int(mes_atual) - 1]['Descrição']

# Processo de Previsão de Entrega
dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].fillna('-')
dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].astype(dtype='str')
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead-Time'].astype(dtype='str')

dfs[int(mes_atual) - 1]['N_Previsão'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'] + '!' + dfs[int(mes_atual) - 1][
    'Lead-Time']

dfs[int(mes_atual) - 1]['N_Previsão'] = dfs[int(mes_atual) - 1]['N_Previsão'].apply(func=lambda val: func(val))

dfs[int(mes_atual) - 1]['Previsão-Entrega'] = dfs[int(mes_atual) - 1]['N_Previsão']

# Processo da Urano
# dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=df_urano, how='left', on='Número', copy=False)
dfs[int(mes_atual) - 1] = pd.merge(left=dfs[int(mes_atual) - 1], right=qy_ent, how='left', on='Número')
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].drop_duplicates(subset=['Número'], keep='first')
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')
# dfs[int(mes_atual) - 1]['Emissão_cte'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'] + '!' + dfs[int(mes_atual) - 1][
#     'Emissão_cte']
dfs[int(mes_atual) - 1]['Data_Ocorrência1'] = dfs[int(mes_atual) - 1]['D_Entrega'] + '!' + dfs[int(mes_atual) - 1][
    'Data_Ocorrência1']
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')
# dfs[int(mes_atual) - 1]['Emissão_cte'] = dfs[int(mes_atual) - 1]['Emissão_cte'].apply(func=lambda val: fatiamento(val))
dfs[int(mes_atual) - 1]['Data_Ocorrência1'] = dfs[int(mes_atual) - 1]['Data_Ocorrência1'].apply(
    func=lambda val: fatiamento(val))
# dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Emissão_cte']
dfs[int(mes_atual) - 1]['D_Entrega'] = dfs[int(mes_atual) - 1]['Data_Ocorrência1']
dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')

# Processo de Dias para Entrega
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead-Time'].apply(func=lambda val: conversor_ldt(val))
dfs[int(mes_atual) - 1]['Previsão-Entrega'] = dfs[int(mes_atual) - 1]['Previsão-Entrega'].astype(dtype='str')
dfs[int(mes_atual) - 1]['D_Entrega'] = dfs[int(mes_atual) - 1]['D_Entrega'].astype(dtype='str')
dfs[int(mes_atual) - 1]['Dias-Para-Entrega'] = dfs[int(mes_atual) - 1]['Previsão-Entrega'] + '!' + \
                                               dfs[int(mes_atual) - 1]['D_Entrega'] + '!' + dfs[int(mes_atual) - 1][
                                                   'Lead-Time']
print("40% concluído")
dfs[int(mes_atual) - 1]['Dias-Para-Entrega'] = dfs[int(mes_atual) - 1]['Dias-Para-Entrega'].apply(
    func=lambda val: dias_p_entrega(val))

# Conversões finais
dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].apply(
    lambda val: to_date_time(val))
dfs[int(mes_atual) - 1]['D_Entrega'] = dfs[int(mes_atual) - 1]['D_Entrega'].apply(
    lambda val: to_date_time(val))
dfs[int(mes_atual) - 1]['Previsão-Entrega'] = dfs[int(mes_atual) - 1]['Previsão-Entrega'].apply(
    lambda val: to_date_time(val))
dfs[int(mes_atual) - 1]['Agendamento'] = dfs[int(mes_atual) - 1]['Agendamento'].apply(
    lambda _: pd.to_datetime(_, format='%d-%m-%Y', errors='coerce'))
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead-Time'].apply(
    func=lambda val: pd.to_numeric(arg=val, errors='coerce'))
dfs[int(mes_atual) - 1]['Lead-Time'] = dfs[int(mes_atual) - 1]['Lead-Time'].fillna('-')

if mes_atual != '1':
    try:

        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'].astype(dtype='str')

        # por algum motivo aqui buga e converte a dfs[int(mes_atual) - 1]['Data-De-Coleta'] para str
        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'].apply(
            func=lambda val: conversor_dt(val))
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2].fillna('-')
        dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=df_col, how='left', on='Número')
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2].fillna('-')
        dfs[int(mes_atual) - 2]['Data-De-Coleta_col'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'] + '!' + \
                                                        dfs[int(mes_atual) - 2]['Data-De-Coleta_col']
        dfs[int(mes_atual) - 2]['Data-De-Coleta_col'] = dfs[int(mes_atual) - 2]['Data-De-Coleta_col'].apply(
            func=lambda val: fatiamento(val))
        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta_col']
        ssw_mes_m = search_data(name_m_ant)  # DF do mês -1
        e_ssw1 = ssw_mes_m.query("Ult_Status == ' MERCADORIA ENTREGUE (01)'")  # DF com apenas as notas entregues
        a_ssw1 = ssw_mes_m.query("Ult_Status == ' ENTREGA AGENDADA (15)'")  # DF com apenas as notas entregues
        a_ssw1['Descrição'] = a_ssw1['Descrição'].apply(func=lambda val: val[19:26] if len(val) > 5 else '')
        e_ssw1['Data'] = e_ssw1['Data'].astype(dtype='str', copy=True)
        e_ssw1['Data'] = e_ssw1['Data'].apply(
            func=lambda val: val[8:12] + '-' + val[5:7] + '-' + val[:4] if len(val) > 3 else val)
        ssw_mes_m = ssw_mes_m[['Número', 'Data', 'Cidade/Estado', 'Ult_Status']]
        ssw_mes_m['Status'] = ssw_mes_m['Ult_Status'] + ' - ' + ssw_mes_m['Cidade/Estado']
        dfs[int(mes_atual) - 2]['Agendamento'] = dfs[int(mes_atual) - 2]['Agendamento'].astype(dtype='str', copy=True)
        dfs[int(mes_atual) - 2]['D_Entrega'] = dfs[int(mes_atual) - 2]['D_Entrega'].astype(dtype='str', copy=True)
        dfs[int(mes_atual) - 2]['D_Entrega'] = dfs[int(mes_atual) - 2]['D_Entrega'].apply(
            func=lambda val: val[8:10] + '-' + val[5:7] + '-' + val[0:4] if len(val) > 3 else val)
        dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=e_ssw1, how='left', on='Número')
        dfs[int(mes_atual) - 2].fillna('-', inplace=True)
        dfs[int(mes_atual) - 2]['Data'] = dfs[int(mes_atual) - 2]['Data'].astype(dtype='str', copy=True)
        dfs[int(mes_atual) - 2]['B_data'] = dfs[int(mes_atual) - 2]['D_Entrega'] + '!' + dfs[int(mes_atual) - 2]['Data']
        dfs[int(mes_atual) - 2]['Data'] = dfs[int(mes_atual) - 2]['B_data'].apply(
            func=lambda val: val[0:11] if len(val) > 11 else val)
        dfs[int(mes_atual) - 2]['Data'] = dfs[int(mes_atual) - 2]['Data'].replace(to_replace='!', value='', regex=True)
        dfs[int(mes_atual) - 2]['D_Entrega'] = dfs[int(mes_atual) - 2]['Data']
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2][all_index]
        dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=ssw_mes_m, how='left')
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2][index_stat]
        dfs[int(mes_atual) - 2].columns = all_index
        dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=a_ssw1, how='left', on='Número')
        dfs[int(mes_atual) - 2]['Descrição'] = dfs[int(mes_atual) - 2]['Descrição'].fillna('')
        dfs[int(mes_atual) - 2]['Agendamento'] = dfs[int(mes_atual) - 2]['Agendamento'].apply(
            func=lambda val: val[0:10] if len(val) > 4 else '')
        print("50% concluído")
        dfs[int(mes_atual) - 2]['Descrição'] = dfs[int(mes_atual) - 2]['Agendamento'] + '!' + dfs[int(mes_atual) - 1][
            'Descrição']
        dfs[int(mes_atual) - 2]['Descrição'] = dfs[int(mes_atual) - 2]['Descrição'].replace(to_replace='!', value='',
                                                                                            regex=True)
        dfs[int(mes_atual) - 2]['Descrição'] = dfs[int(mes_atual) - 2]['Descrição'].astype(dtype='str', copy=True)
        dfs[int(mes_atual) - 2]['Descrição'] = dfs[int(mes_atual) - 2]['Descrição'].apply(
            func=lambda val: val[-6:-4] + '-' + val[-4:-2] + '-' + '20' + val[-2:] if len(val) > 1 else '-')
        dfs[int(mes_atual) - 2]['Agendamento'] = dfs[int(mes_atual) - 2]['Descrição']
        dfs[int(mes_atual) - 2]['Agendamento'] = dfs[int(mes_atual) - 2]['Agendamento'].apply(
            lambda _: pd.to_datetime(_, format='%d-%m-%Y', errors='coerce'))
        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'].fillna('-')
        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'].astype(dtype='str')
        dfs[int(mes_atual) - 2]['Lead-Time'] = dfs[int(mes_atual) - 2]['Lead-Time'].astype(dtype='str')
        dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=df_urano, how='left', on='Número',
                                           copy=False)
        dfs[int(mes_atual) - 2] = pd.merge(left=dfs[int(mes_atual) - 2], right=qy_ent, how='left', on='Número')
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2].drop_duplicates(subset=['Número'], keep='first')
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2].fillna('-')
        dfs[int(mes_atual) - 2]['Emissão_cte'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'] + '!' + \
                                                 dfs[int(mes_atual) - 2][
                                                     'Emissão_cte']
        dfs[int(mes_atual) - 2]['Data_Ocorrência1'] = dfs[int(mes_atual) - 2]['D_Entrega'] + '!' + \
                                                      dfs[int(mes_atual) - 2][
                                                          'Data_Ocorrência1']
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2].fillna('-')
        dfs[int(mes_atual) - 2]['Emissão_cte'] = dfs[int(mes_atual) - 2]['Emissão_cte'].apply(
            func=lambda val: fatiamento(val))
        dfs[int(mes_atual) - 2]['Data_Ocorrência1'] = dfs[int(mes_atual) - 2]['Data_Ocorrência1'].apply(
            func=lambda val: fatiamento(val))
        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Emissão_cte']
        dfs[int(mes_atual) - 2]['D_Entrega'] = dfs[int(mes_atual) - 2]['Data_Ocorrência1']
        dfs[int(mes_atual) - 2] = dfs[int(mes_atual) - 2].fillna('-')
        dfs[int(mes_atual) - 2]['Lead-Time'] = dfs[int(mes_atual) - 2]['Lead-Time'].apply(
            func=lambda val: conversor_ldt(val))
        dfs[int(mes_atual) - 2]['Previsão-Entrega'] = dfs[int(mes_atual) - 2]['Previsão-Entrega'].astype(dtype='str')
        dfs[int(mes_atual) - 2]['D_Entrega'] = dfs[int(mes_atual) - 2]['D_Entrega'].astype(dtype='str')
        dfs[int(mes_atual) - 2]['Dias-Para-Entrega'] = dfs[int(mes_atual) - 2]['Previsão-Entrega'] + '!' + \
                                                       dfs[int(mes_atual) - 2]['D_Entrega'] + '!' + \
                                                       dfs[int(mes_atual) - 2][
                                                           'Lead-Time']
        dfs[int(mes_atual) - 2]['Dias-Para-Entrega'] = dfs[int(mes_atual) - 2]['Dias-Para-Entrega'].apply(
            func=lambda val: dias_p_entrega(val))
        dfs[int(mes_atual) - 2]['Data-De-Coleta'] = dfs[int(mes_atual) - 2]['Data-De-Coleta'].apply(
            lambda _: pd.to_datetime(_, format='%d-%m-%Y', errors='coerce'))
        dfs[int(mes_atual) - 2]['D_Entrega'] = dfs[int(mes_atual) - 2]['D_Entrega'].apply(
            lambda _: pd.to_datetime(_, format='%d-%m-%Y', errors='coerce'))
        dfs[int(mes_atual) - 2]['Lead-Time'] = dfs[int(mes_atual) - 2]['Lead-Time'].apply(
            func=lambda val: pd.to_numeric(arg=val, errors='coerce'))
        dfs[int(mes_atual) - 2]['Lead-Time'] = dfs[int(mes_atual) - 2]['Lead-Time'].fillna('-')
    except NameError:
        # Exceção já tratada no IF
        pass

print("60% concluído")

dfs[int(mes_atual) - 1]['Data-De-Coleta'] = dfs[int(mes_atual) - 1]['Data-De-Coleta'].apply(
    lambda val: to_date_time(val))

dfs[int(mes_atual) - 1] = dfs[int(mes_atual) - 1].fillna('-')

############################
cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
dates = ['C', 'L', 'M', 'N', 'O']

book = Workbook()

index_style = NamedStyle(name='index_style',
                         fill=PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid'))
index_style.font = Font(size=9, bold=True)
bd_index = Side(style='thin', color='000000')
index_style.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
index_style.alignment = Alignment(horizontal='center')
book.add_named_style(index_style)

date_format = NamedStyle(name='date_format', number_format='DD/MM/YYYY', font=Font(color='000000'))
date_format.font = Font(size=8)
date_format.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
date_format.alignment = Alignment(horizontal='center')
book.add_named_style(date_format)

default_style = NamedStyle(name='default_style', font=Font(color='000000'))
default_style.font = Font(size=8)
default_style.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
default_style.alignment = Alignment(horizontal='center')
book.add_named_style(default_style)
print("70% concluído")
try:
    for x in range(1, len(dfs) + 1):
        sheet_name = meses_str[str(x)] + '-' + ano_atual
        print(sheet_name + ' incluso na planilha')

        if x == 1:
            date = NamedStyle(name='date', number_format='DD/MM/YYYY', font=Font(color='000000'))
            date.font = Font(size=8)
            date.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
            date.alignment = Alignment(horizontal='center')
            book.add_named_style(date)

        if x > 1:
            book = load_workbook(filename=dest_file)
            writer = pd.ExcelWriter(path=dest_file, engine='openpyxl')
            writer.book = book

        if x == 1:
            writer = pd.ExcelWriter(path=dest_file,
                                    engine='openpyxl', index=False)

        dfs[x - 1] = dfs[x - 1][all_index]

        dfs[x - 1].to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]

        for z in cols:
            for cell in ws[z]:
                if cell not in dates:
                    cell.style = default_style

            if z == 'A' or 'B':
                ws.column_dimensions[z].width = 9

            if z in dates:
                ws.column_dimensions[z].width = 13
                for cell in ws[z]:
                    cell.style = date_format

                if x >= 2:
                    for cell in ws[z]:
                        cell.style = date

            if z == 'D':
                ws.column_dimensions[z].width = 33

            if z == 'E':
                ws.column_dimensions[z].width = 30

            if z == 'F':
                ws.column_dimensions[z].width = 4

            if z == 'G':
                ws.column_dimensions[z].width = 24

            if z == 'H':
                ws.column_dimensions[z].width = 13

            if z == 'I':
                ws.column_dimensions[z].width = 37

            if z == 'J':
                ws.column_dimensions[z].width = 25

            if z == 'K':
                ws.column_dimensions[z].width = 25

            if z == 'Q':
                ws.column_dimensions[z].width = 14

            if z == 'S':
                ws.column_dimensions[z].width = 48

            # Index
            col = z + '1'
            ws[col].style = index_style

        writer.save()
except IndexError:
    writer.save()
print("80% concluído")
# Clientes com Agendamento

book = load_workbook(filename=dest_file)
writer = pd.ExcelWriter(path=dest_file, engine='openpyxl')
writer.book = book
itens_comp[0].to_excel(excel_writer=writer, sheet_name=sheet_names[0], index=False)

ws = writer.sheets[sheet_names[0]]

style2 = NamedStyle(name='style2', font=Font(size=10, color='000000'))
style2.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
style2.alignment = Alignment(horizontal='center', vertical='center')
book.add_named_style(style2)

for c in cols[0:3]:
    for cell in ws[c]:
        cell.style = style2

    if c == 'A':
        ws.column_dimensions[c].width = 45

    if c == 'B':
        ws.column_dimensions[c].width = 75

    if c == 'C':
        ws.column_dimensions[c].width = 70

    # Index
    col = c + '1'
    ws[col].style = index_style
print("90% concluído")
ws.merge_cells('A4:A5')
ws.merge_cells('A6:A7')
ws.merge_cells('A8:A13')
ws.merge_cells('A29:A30')
ws.merge_cells('A34:A35')

print(sheet_names[0] + ' Incluso na Planilha')

writer.save()

# Contato Transportadora

book = load_workbook(filename=dest_file)
writer = pd.ExcelWriter(path=dest_file, engine='openpyxl')
writer.book = book
itens_comp[1].to_excel(excel_writer=writer, sheet_name=sheet_names[1], index=False)

ws = writer.sheets[sheet_names[1]]

style3 = NamedStyle(name='style3', font=Font(size=10, color='000000'))
style3.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
style3.alignment = Alignment(horizontal='center', vertical='center')
book.add_named_style(style3)

for c in cols[0:5]:
    for cell in ws[c]:
        cell.style = style3

    if c == 'A':
        ws.column_dimensions[c].width = 18

    if c == 'B':
        ws.column_dimensions[c].width = 25

    if c == 'C':
        ws.column_dimensions[c].width = 40

    if c == 'D':
        ws.column_dimensions[c].width = 45

    if c == 'E':
        ws.column_dimensions[c].width = 65
    # Index
    col = c + '1'
    ws[col].style = index_style

ws.merge_cells('A2:A6')
ws.merge_cells('A7:A8')
ws.merge_cells('A9:A13')
ws.merge_cells('A14:A18')
ws.merge_cells('A19:A23')
ws.merge_cells('A24:A29')
ws.merge_cells('A30:A32')
ws.merge_cells('A33:A35')
ws.merge_cells('A36:A38')
ws.merge_cells('A39:A44')
ws.merge_cells('A45:A47')
ws.merge_cells('A48:A56')
ws.merge_cells('A57:A64')
ws.merge_cells('A65:A70')
ws.merge_cells('A72:A75')
ws.merge_cells('A76:A80')
ws.merge_cells('A81:A82')
ws.merge_cells('A83:A89')
ws.merge_cells('A90:A95')
ws.merge_cells('A96:A97')
ws.merge_cells('A98:A104')
ws.merge_cells('A105:A106')
ws.merge_cells('B2:B6')
ws.merge_cells('B7:B8')
ws.merge_cells('B9:B13')
ws.merge_cells('B14:B18')
ws.merge_cells('B19:B23')
ws.merge_cells('B24:B29')
ws.merge_cells('B30:B32')
ws.merge_cells('B33:B35')
ws.merge_cells('B36:B38')
ws.merge_cells('B39:B44')
ws.merge_cells('B45:B47')
ws.merge_cells('B48:B56')
ws.merge_cells('B57:B64')
ws.merge_cells('B65:B70')
ws.merge_cells('B72:B75')
ws.merge_cells('B76:B80')
ws.merge_cells('B81:B82')
ws.merge_cells('B83:B89')
ws.merge_cells('B90:B95')
ws.merge_cells('B96:B97')
ws.merge_cells('B98:B104')
ws.merge_cells('B105:B106')

print(sheet_names[1] + ' Incluso na Planilha')

writer.save()
print("--- %s seconds creation of plan ---" % (time.time() - start_time))
