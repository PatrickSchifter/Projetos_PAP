import pandas as pd
from Projetos.Rotina.config import index_stat, all_index, dest_file, dia_atual, ano_atual, mes_atual, dia_semana, func, file, n_file, \
    meses_str, manual_index, aut_index, fatiamento, conversor_dt, conversor_ldt, dias_p_entrega, \
    to_date_time, aut_index_a
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle, Font, Border, Side, PatternFill
from openpyxl import Workbook
import time
import os
import shutil


itens_comp = []
sheet_names = []

sheet_name = 'Clientes com Agendamento'
df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
sheet_names.append(sheet_name)
itens_comp.append(df)

sheet_name = 'Contato Transportadoras'
df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
sheet_names.append(sheet_name)
itens_comp.append(df)

df_n = pd.read_excel(n_file, engine='openpyxl')
df_n.columns = aut_index_a
df_n['1'] = pd.NA
df_n['2'] = pd.NA
df_n['3'] = pd.NA
df_n['4'] = pd.NA
df_n['5'] = pd.NA
df_n['6'] = pd.NA
df_n['7'] = pd.NA
df_n['8'] = pd.NA

df_n = df_n.fillna('-')


df_n.columns = all_index
print(df_n.columns)
dfs = []

try:
    for dfss in range(1, 13):
        dfss = str(dfss)
        sheet = meses_str[dfss] + '-' + ano_atual
        try:
            df = pd.read_excel(file, engine='openpyxl', sheet_name=sheet)
            df.columns = all_index
        except ValueError:
            continue

        dfs.append(df)
except IndexError:
    pass


dfs.append(df_n)

cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
dates = ['C', 'L', 'M', 'N', 'O']

book = Workbook()

index_style = NamedStyle(name='index_style',
                         fill=PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid'))
index_style.font = Font(size=9, bold=True)
bd_index = Side(style='thin', color='000000')
index_style.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
index_style.alignment = Alignment(horizontal='center')
book.add_named_style(index_style)

date_format = NamedStyle(name='date_format', number_format='DD/MM/YYYY', font=Font(color='000000'))
date_format.font = Font(size=8)
date_format.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
date_format.alignment = Alignment(horizontal='center')
book.add_named_style(date_format)

default_style = NamedStyle(name='default_style', font=Font(color='000000'))
default_style.font = Font(size=8)
default_style.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
default_style.alignment = Alignment(horizontal='center')
book.add_named_style(default_style)
print("70% concluído")
try:
    for x in range(1, len(dfs) + 1):
        sheet_name = meses_str[str(x)] + '-' + ano_atual
        print(sheet_name + ' incluso na planilha')

        if x == 1:
            date = NamedStyle(name='date', number_format='DD/MM/YYYY', font=Font(color='000000'))
            date.font = Font(size=8)
            date.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
            date.alignment = Alignment(horizontal='center')
            book.add_named_style(date)

        if x > 1:
            book = load_workbook(filename=dest_file)
            writer = pd.ExcelWriter(path=dest_file, engine='openpyxl')
            writer.book = book

        if x == 1:
            writer = pd.ExcelWriter(path=dest_file,
                                    engine='openpyxl', index=False)

        dfs[x - 1] = dfs[x - 1][all_index]

        dfs[x - 1].to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]

        for z in cols:
            for cell in ws[z]:
                if cell not in dates:
                    cell.style = default_style

            if z == 'A' or 'B':
                ws.column_dimensions[z].width = 9

            if z in dates:
                ws.column_dimensions[z].width = 13
                for cell in ws[z]:
                    cell.style = date_format

                if x >= 2:
                    for cell in ws[z]:
                        cell.style = date

            if z == 'D':
                ws.column_dimensions[z].width = 33

            if z == 'E':
                ws.column_dimensions[z].width = 30

            if z == 'F':
                ws.column_dimensions[z].width = 4

            if z == 'G':
                ws.column_dimensions[z].width = 24

            if z == 'H':
                ws.column_dimensions[z].width = 13

            if z == 'I':
                ws.column_dimensions[z].width = 37

            if z == 'J':
                ws.column_dimensions[z].width = 25

            if z == 'K':
                ws.column_dimensions[z].width = 25

            if z == 'Q':
                ws.column_dimensions[z].width = 14

            if z == 'S':
                ws.column_dimensions[z].width = 48

            # Index
            col = z + '1'
            ws[col].style = index_style

        writer.save()
except IndexError:
    writer.save()
print("80% concluído")
# Clientes com Agendamento

book = load_workbook(filename=dest_file)
writer = pd.ExcelWriter(path=dest_file, engine='openpyxl')
writer.book = book
itens_comp[0].to_excel(excel_writer=writer, sheet_name=sheet_names[0], index=False)

ws = writer.sheets[sheet_names[0]]

style2 = NamedStyle(name='style2', font=Font(size=10, color='000000'))
style2.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
style2.alignment = Alignment(horizontal='center', vertical='center')
book.add_named_style(style2)

for c in cols[0:3]:
    for cell in ws[c]:
        cell.style = style2

    if c == 'A':
        ws.column_dimensions[c].width = 45

    if c == 'B':
        ws.column_dimensions[c].width = 75

    if c == 'C':
        ws.column_dimensions[c].width = 70

    # Index
    col = c + '1'
    ws[col].style = index_style
print("90% concluído")
ws.merge_cells('A4:A5')
ws.merge_cells('A6:A7')
ws.merge_cells('A8:A13')
ws.merge_cells('A29:A30')
ws.merge_cells('A34:A35')

print(sheet_names[0] + ' Incluso na Planilha')

writer.save()

# Contato Transportadora

book = load_workbook(filename=dest_file)
writer = pd.ExcelWriter(path=dest_file, engine='openpyxl')
writer.book = book
itens_comp[1].to_excel(excel_writer=writer, sheet_name=sheet_names[1], index=False)

ws = writer.sheets[sheet_names[1]]

style3 = NamedStyle(name='style3', font=Font(size=10, color='000000'))
style3.border = Border(left=bd_index, right=bd_index, top=bd_index, bottom=bd_index)
style3.alignment = Alignment(horizontal='center', vertical='center')
book.add_named_style(style3)

for c in cols[0:5]:
    for cell in ws[c]:
        cell.style = style3

    if c == 'A':
        ws.column_dimensions[c].width = 18

    if c == 'B':
        ws.column_dimensions[c].width = 25

    if c == 'C':
        ws.column_dimensions[c].width = 40

    if c == 'D':
        ws.column_dimensions[c].width = 45

    if c == 'E':
        ws.column_dimensions[c].width = 65
    # Index
    col = c + '1'
    ws[col].style = index_style

ws.merge_cells('A2:A6')
ws.merge_cells('A7:A8')
ws.merge_cells('A9:A13')
ws.merge_cells('A14:A18')
ws.merge_cells('A19:A23')
ws.merge_cells('A24:A29')
ws.merge_cells('A30:A32')
ws.merge_cells('A33:A35')
ws.merge_cells('A36:A38')
ws.merge_cells('A39:A44')
ws.merge_cells('A45:A47')
ws.merge_cells('A48:A56')
ws.merge_cells('A57:A64')
ws.merge_cells('A65:A70')
ws.merge_cells('A72:A75')
ws.merge_cells('A76:A80')
ws.merge_cells('A81:A82')
ws.merge_cells('A83:A89')
ws.merge_cells('A90:A95')
ws.merge_cells('A96:A97')
ws.merge_cells('A98:A104')
ws.merge_cells('A105:A106')
ws.merge_cells('B2:B6')
ws.merge_cells('B7:B8')
ws.merge_cells('B9:B13')
ws.merge_cells('B14:B18')
ws.merge_cells('B19:B23')
ws.merge_cells('B24:B29')
ws.merge_cells('B30:B32')
ws.merge_cells('B33:B35')
ws.merge_cells('B36:B38')
ws.merge_cells('B39:B44')
ws.merge_cells('B45:B47')
ws.merge_cells('B48:B56')
ws.merge_cells('B57:B64')
ws.merge_cells('B65:B70')
ws.merge_cells('B72:B75')
ws.merge_cells('B76:B80')
ws.merge_cells('B81:B82')
ws.merge_cells('B83:B89')
ws.merge_cells('B90:B95')
ws.merge_cells('B96:B97')
ws.merge_cells('B98:B104')
ws.merge_cells('B105:B106')

print(sheet_names[1] + ' Incluso na Planilha')

writer.save()
value = True
while value:
    try:
        os.remove(file)
        shutil.copy(src=dest_file, dst=file)
        value = False
    except PermissionError:
        import time
        import win32com.client as win32

        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'aline.gomes@portoaporto.com.br; luana.neves@portoaporto.com.br'
        mail.Subject = 'Mensagem do Robô'
        mail.Body = 'Olá pessoal, aqui é o robô.\n\nEstou tentando excluir o arquivo de monitoramento para colocar um novinho em folha. Preciso que quem está com ele aberto o feche para que eu possa terminar meu trabalho.\n\nMuito obrigado'
        mail.Send()
        time.sleep(300)
